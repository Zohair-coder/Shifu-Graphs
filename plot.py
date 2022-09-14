from openpyxl import Workbook, load_workbook
import os
import re
import csv
import matplotlib.pyplot as plt

ROOT_DIR = "Archive"


def main():
    data = get_data_from_files()
    create_graphs(data)


def get_data_from_files():
    all_data = {}
    directories = os.listdir(ROOT_DIR)
    for directory in directories:
        if is_unnecessary_directory(directory):
            continue
        data = get_data_from_ff_directory(directory)
        all_data[directory] = data
    return all_data


def create_graphs(data):
    for dir_name, ff_data in data.items():
        fig, ax = plt.subplots()
        ax.set_xlabel("Extension (mm)")
        ax.set_ylabel("Load (N)")
        ax.set_title(dir_name)
        ax.grid(True)
        for specimen_number, specimen_data in ff_data.items():
            x = []
            y = []
            for data_point in specimen_data:
                x.append(data_point["extension"])
                y.append(data_point["load"])
            ax.plot(x, y, label="Specimen " + specimen_number)
        ax.legend()
        file_location = "Graphs/" + dir_name + ".png"
        if not os.path.exists(os.path.dirname(file_location)):
            os.makedirs(os.path.dirname(file_location))
        plt.savefig(file_location)
        plt.close()


def get_data_from_ff_directory(directory):
    all_data = {}
    ff_dir_path = os.path.join(ROOT_DIR, directory)
    ff_dir_contents = os.listdir(ff_dir_path)
    for content in ff_dir_contents:
        final_dir_path = os.path.join(ff_dir_path, content)
        if os.path.isdir(final_dir_path):
            final_dir_contents = os.listdir(final_dir_path)
            for file in final_dir_contents:
                convert_csv_to_excel(final_dir_path, file)
            final_dir_contents = os.listdir(final_dir_path)
            for file in final_dir_contents:
                final_file_path = os.path.join(
                    final_dir_path, file)
                if file.endswith(".xlsx"):
                    data = get_data_from_file(final_file_path)
                    specimen_number = get_specimen_number(file)
                    all_data[specimen_number] = data
    return all_data


def convert_csv_to_excel(final_dir_path, file):
    if file.endswith(".csv"):
        convert_to_excel(os.path.join(final_dir_path, file))
        os.remove(os.path.join(final_dir_path, file))


def get_data_row(sheet):
    for row in sheet.iter_rows():
        if row[0].value == "Time":
            return row[0].row + 2


def get_data_from_file(file):
    workbook = load_workbook(file, read_only=True, data_only=True)
    sheet = workbook.active
    data = []
    data_row = get_data_row(sheet)
    for row in sheet.iter_rows(min_row=data_row):
        time = float(row[0].value)
        extension = float(row[1].value)
        load = float(row[2].value)
        data.append({
            "time": time,
            "extension": extension,
            "load": load
        })
    return data


def get_specimen_number(file):
    matches = re.search(
        r'Specimen_RawData_(\d+).xlsx', file)
    specimen_number = matches.group(1)
    return specimen_number


def is_unnecessary_directory(directory):
    return directory == ".DS_Store"


def convert_to_excel(path):
    wb = Workbook()
    ws = wb.active
    with open(path) as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            ws.append(row)
    wb.save(filename=path.replace(".csv", ".xlsx"))


if __name__ == "__main__":
    main()
